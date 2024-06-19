import asyncio
import logging
import time
from datetime import datetime
from pathlib import Path
from multiprocessing import cpu_count, Process

from openpyxl import Workbook, load_workbook

from config import DATE_FORMAT
from src.my_logging import get_logger
from src.scrappers.nopriz import ScraperNopriz
from src.scrappers.nostroy import ScraperNostroy
from src.schemas import NostroyRow, NoprizRow, FiltersNostroy, FiltersNopriz

DATE_FROM = '01.01.1900'
# DATE_FROM = '13.06.2024'
DATE_TO = datetime.now().date().strftime(DATE_FORMAT)
date_from = datetime.strptime(DATE_FROM, DATE_FORMAT)
date_to = datetime.strptime(DATE_TO, DATE_FORMAT)


def write_data_to_excel(filepath: Path, data: list[NoprizRow | NostroyRow]) -> None:
    if filepath.exists():
        wb = load_workbook(filepath)
        ws = wb.worksheets[0]
    else:
        wb = Workbook()
        ws = wb.worksheets[0]
        columns = list(data[0].dict(by_alias=True).keys())
        ws.append(columns)

    i = 0
    while True:
        try:
            for r in data:
                d = r.dict(by_alias=True)
                ws.append(list(d.values()))

            wb.save(filepath)
            logging.info(f'{filepath} was written')
            break
        except PermissionError as ex:
            logging.info('Please close the file')
            i += 1
            time.sleep(10)
            if i == 500:
                raise ex


def run_scrapper(filepath: Path, ids: list[int]):
    asyncio.run(_run_scrapper(filepath=filepath, ids=ids))


async def _run_scrapper(filepath: Path, ids: list[int]):
    data_to_append = []
    async with ScraperNopriz(date_format=DATE_FORMAT, date_from=date_from, date_to=date_to) as scrapper:
        async for data in scrapper.collect_data(ids):
            for row in data:
                data_to_append.append(row)

            if len(data_to_append) >= 1000:
                logging.info(f'Appending {data_to_append} collected data to excel')
                scrapper.to_excel(filepath=filepath, data=data_to_append)
                logging.info('End of appending data')
                data_to_append = []

        if data_to_append:
            scrapper.to_excel(filepath, data_to_append)


async def main():
    proxy_url = ''

    async with ScraperNopriz(date_format=DATE_FORMAT, date_from=date_from, date_to=date_to) as scrapper:
        filters = FiltersNopriz(member_status=1, sro_enabled=True, sro_registration_number='П')
        ids = await scrapper.get_ids(filters=filters, use_cached=True)

    ids = ids[7000:]
    logging.info(f'{len(ids)=}')

    count_processes = cpu_count()
    count_ids_per_task = len(ids) // count_processes
    ids_for_tasks = [ids[i:i+count_ids_per_task] for i in range(0, len(ids), count_ids_per_task)]

    for i in range(count_processes):
        filename = scrapper.get_filename('nopriz') + f'_proektirovshiki_{i}'
        filepath = Path(filename).with_suffix('.xlsx')

        p = Process(target=run_scrapper, args=(filepath, ids_for_tasks[i]))
        p.start()


    # async with ScraperNostroy(proxy_url=proxy_url, date_format=DATE_FORMAT, date_from=date_from, date_to=date_to) as scrapper:
    #     filters = FiltersNostroy(member_status=1, sro_enabled=True)
    #     ids = await scrapper.get_ids(filters=filters)
    #     filepath = Path(scrapper.get_filename('nostroy')).with_suffix('.xlsx')
    #
    #     data_to_append = []
    #     async for data in scrapper.collect_data(ids):
    #         for row in data:
    #             data_to_append.append(row)
    #
    #         if len(data_to_append) >= 10000:
    #             write_data_to_excel(filepath=filepath, data=data_to_append)
    #             data_to_append = []
    #
    #     if data_to_append:
    #         write_data_to_excel(filepath=filepath, data=data_to_append)


if __name__ == '__main__':
    get_logger(Path(__file__).stem + '.log')

    try:
        asyncio.run(main())
    except Exception as ex:
        logging.error(ex, exc_info=True)
        exit(1)
