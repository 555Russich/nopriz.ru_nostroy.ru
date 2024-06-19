import logging
import re
from datetime import datetime
from pathlib import Path

import aiofiles
from openpyxl import load_workbook, Workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side

from src.scrappers.base import BaseScrapper
from src.schemas import NoprizRow, SRO, Check, Insurance


class ScraperNopriz(BaseScrapper):
    DATE_API_FORMAT = '%Y-%m-%dT%X%z'
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                             '(KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'}

    async def _collect_ids(self, filters: dict):
        ids = set()
        collect_per_page = 1000
        json_ = {
            'filters': filters,
            'page': 1,
            'pageCount': f"{collect_per_page}",
            'searchString': "",
            'sortBy': {
                'registry_registration_date': 'DESC'
            }
        }
        while True:
            r_json = await self.request_json(
                    method='POST',
                    url='https://reestr.nopriz.ru/api/sro/all/member/list',
                    headers=self.headers,
                    proxy=self.proxy_url,
                    json=json_
            )

            for d in r_json['data']['data']:
                dt = datetime.strptime(
                    re.sub(r'\+0\d:00', '', d['registry_registration_date']),
                    '%Y-%m-%dT%H:%M:%S'
                )
                if dt < self.date_from:
                    logging.info(f'Page №: {json_["page"]}, last date: {dt}')
                    logging.info(f'{len(ids)} ids was collected')
                    return list(ids)
                elif dt > self.date_to:
                    continue
                else:
                    ids.add(d['id'])
            else:
                logging.info(f'Page №: {json_["page"]}, last date: {dt}')
                logging.info(f'Page №: {json_["page"]} ; Collected ids: {len(ids)}')

            if len(r_json['data']['data']) < collect_per_page:
                break
            json_["page"] += 1
        return list(ids)

    async def get_sro_member(self, id_: int) -> NoprizRow:
        r = await self.request_json(
            method='POST',
            url=f'https://reestr.nopriz.ru/api/member/{id_}/info',
            headers=self.headers,
            proxy=self.proxy_url
        )
        r = r['data']

        sro_id = int(r['sro']['id'])
        if sro_id in self.sro_cache:
            sro = self.sro_cache[sro_id]
        else:
            sro = await self.get_sro(sro_id)
            self.sro_cache[id_] = sro

        is_simple = r['member_right_vv']['is_simple'] if r.get('member_right_vv') else False
        is_simple = 'Да' if is_simple else 'Нет'
        is_nuclear = r['member_right_vv']['is_nuclear'] if r.get('member_right_vv') else False
        is_nuclear = 'Да' if is_nuclear else 'Нет'
        is_extremely_dangerous = r['member_right_vv']['is_extremely_dangerous'] if r.get('member_right_vv') else False
        is_extremely_dangerous = 'Да' if is_extremely_dangerous else 'Нет'

        checks = [Check(
                    check_date=datetime.strptime(c['check_date'], self.DATE_API_FORMAT).strftime(self.date_format)
                    if c.get('check_date') else None,
                    disciplinary_action=c['disciplinary_action']['title'] if c.get('disciplinary_action') else None,
                    member_check_result=c['member_check_result']['title'] if c.get('member_check_result') else None,
                    member_check_type=c['member_check_type']['title'] if c.get('member_check_type') else None,
        ) for c in r.get('checks', [])]

        insurances = [Insurance(
            begin_date=datetime.strptime(i['begin_date'], self.DATE_API_FORMAT).strftime(self.date_format)
            if i.get('begin_date') else None,
            contract_number=i.get('contract_number'),
            end_date=datetime.strptime(i['end_date'], self.DATE_API_FORMAT).strftime(self.date_format)
            if i.get('end_date') else None,
            insurance_sum=i.get('insurance_summ'),
            insurer=i.get('insurer'),
            license=i.get('license'),
            object_title=i['object']['title'] if i.get('object') else None,
            phone=i.get('phone'),
            place=i.get('place'),
        ) for i in r.get('insurances', [])]

        try:
            return NoprizRow(
                id=r['id'],
                sro=sro,
                registration_number=r['registration_number'],
                full_description=r.get('full_description'),
                short_description=r.get('short_description'),
                member_type=r['member_type'].get('title') if r.get('member_type') else None,
                registry_registration_date=datetime.strptime(
                    r['registry_registration_date'], self.DATE_API_FORMAT
                ).strftime(f'{self.date_format}') if r.get('registry_registration_date') else None,
                basis=r.get('basis'),
                approved_basis_date=datetime.strptime(
                    r['approved_basis_date'], self.DATE_API_FORMAT
                ).strftime(f'{self.date_format}') if r.get('approved_basis_date') else None,
                ogrnip=r.get('ogrnip'),
                inn=r.get('inn'),
                phones=r.get('phones'),
                full_address=NoprizRow.parse_full_address(r),
                director=r.get('director'),
                member_status=r['member_status'].get('title') if r.get('member_status') else None,
                accordance_status=r['accordance_status'].get('title') if r.get('accordance_status') else None,
                created_at=datetime.strptime(
                    r['created_at'], self.DATE_API_FORMAT
                ).strftime(f'{self.date_format}') if r.get('created_at') else None,
                last_updated_at=datetime.strptime(
                    r['last_updated_at'], self.DATE_API_FORMAT
                ).strftime(f'{self.date_format} %X') if r.get('last_updated_at') else None,
                suspension_date=datetime.strptime(
                    r['suspension_date'], self.DATE_API_FORMAT
                ).strftime(f'{self.date_format} %X') if r.get('suspension_date') else None,
                suspension_reason=r.get('suspension_reason'),
                member_right_vv=r['member_right_vv'].get('compensation_fund') if r.get('member_right_vv') else None,
                member_right_odo=r['member_right_odo'].get('compensation_fund') if r.get('member_right_odo') else None,
                right_status=r['right']['right_status']['title'] if r.get('right') else None,
                right_status_vv=r['member_right_vv']['right_status']['title'] if r.get('member_right_vv') else None,
                right_status_odo=r['member_right_odo']['right_status']['title'] if r.get('member_right_odo') else None,
                odo_compensation_fund_date=datetime.strptime(
                    r['member_right_odo']['right_status_date'], self.DATE_API_FORMAT
                ).strftime(self.date_format) if r.get('member_right_odo') and
                                                r['member_right_odo'].get('right_status_date') else None,
                odo_responsibility_level_date=datetime.strptime(
                    r['member_right_odo']['responsibility_level_date'], self.DATE_API_FORMAT
                ).strftime(self.date_format) if r.get('member_right_odo') and
                                                r['member_right_odo'].get('responsibility_level_date') else None,
                right_basis=r['right']['basis'] if r.get('right') else None,
                is_simple=is_simple,
                simple_date=datetime.strptime(
                    r['member_right_vv']['simple_date'], self.DATE_API_FORMAT
                ).strftime(self.date_format)
                if r.get('member_right_vv') and r['member_right_vv'].get('simple_date') else None,
                is_extremely_dangerous=is_extremely_dangerous,
                extremely_dangerous_date=datetime.strptime(
                    r['member_right_vv']['extremely_dangerous_date'], self.DATE_API_FORMAT
                ).strftime(self.date_format)
                if r.get('member_right_vv') and r['member_right_vv'].get('extremely_dangerous_date') else None,
                is_nuclear=is_nuclear,
                nuclear_date=datetime.strptime(
                    r['member_right_vv']['nuclear_date'], self.DATE_API_FORMAT
                ).strftime(self.date_format)
                if r.get('member_right_vv') and r['member_right_vv'].get('nuclear_date') else None,
                responsibility_level_odo=f'{r['member_right_odo']['responsibility_level']['title']}, '
                                         f'{r['member_right_odo']['responsibility_level']['cost']}'
                if r.get('member_right_odo') and r['member_right_odo'].get('responsibility_level') else None,
                responsibility_level_vv=f'{r['member_right_vv']['responsibility_level']['title']}, '
                                        f'{r['member_right_vv']['responsibility_level']['cost']}'
                if r.get('member_right_vv') and r['member_right_vv'].get('responsibility_level') else None,
                checks=checks,
                insurances=insurances
            )
        except Exception as ex:
            logging.info(f'{id_=}\n{r}')
            raise ex

    async def get_sro(self, id_: int) -> SRO:
        d = await self.request_json(method='POST', url=f'https://reestr.nopriz.ru/api/sro/{id_}')
        d = d['data']
        return SRO(
            full_description=d['full_description'],
            short_description=d['short_description'],
            registration_number=d['registration_number'],
            inn=d['inn'],
            ogrn=d['ogrn'],
            place=d['place'],
            phone=d['phone'],
            email=d['email'],
            site=d['site']
        )

    async def download_xlsx_cart(self, id_: int) -> bytes:
        async with self._session.request(
                method='GET',
                url=f'https://reestr.nopriz.ru/api/member/{id_}/cart/download',
                timeout=30
        ) as r:
            return await r.read()

    @classmethod
    def to_excel(cls, filepath: Path, data: list[NoprizRow]) -> None:
        if filepath.exists():
            wb = load_workbook(filepath)
            ws = wb.worksheets[0]
        else:
            wb = Workbook()
            ws = wb.worksheets[0]

            first_row = {
                'Сведения о СРО': 'A1',
                'Сведения о члене саморегулируемой организации': 'B1:P1',
                'Размер обязательств': 'Q1:R1',
                'Сведения о КФ': 'S1:V1',
                'Сведения о наличии права': 'W1:AE1',
                'Сведения о проверках': 'AF1:AI1',
                'Сведения об обеспечении имущественной ответственности': 'AJ1:AN1'
            }
            thin = Side(border_style='thin', color='000000')
            for v, range_string in first_row.items():
                first_cell = range_string.split(':')[0]
                ws[first_cell] = v
                ws[first_cell].font = Font(bold=True, size=10)
                ws[first_cell].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.merge_cells(range_string=range_string)

            second_row = [
                'Полное и (в случае, если имеется) сокращенное наименование саморегулируемой организации, основной государственный регистрационный номер записи о государственной регистрации юридического лица, идентификационный номер налогоплательщика, регистрационный номер в государственном реестре саморегулируемых организаций, адрес места нахождения, адрес официального сайта в информационно-телекоммуникационной сети "Интернет", адрес электронной почты, контактный телефон',
                'N п/п',
                'Полное наименование',
                'Сокращенное наименование',
                'Регистрационный номер в реестре СРО',
                'Дата регистрации в реестре СРО',
                # 'Дата государственной регистрации',
                'Статус (члена СРО)',
                'Дата прекращения членства',
                'Основание прекращения членства',
                'ОГРН/ОГРНИП',
                'ИНН',
                'Контактные телефоны',
                'Адрес места нахождения юридического лица',
                'Фамилия, имя, отчество лица, осуществляющего функции единоличного исполнительного органа юридического лица и (или) руководителя коллегиального исполнительного органа юридического лица',
                # 'Фамилия, имя, отчество (при наличии) для ИП',
                # 'Адрес места фактического осуществления деятельности для ИП',
                'Сведения о соответствии',
                'Иные сведения, предусмотренные требованиями СРО',
                'Фактический совокупный размер обязательств члена саморегулируемой организации по договорам подряда',
                'Дата обновления',
                'Размер взноса в компенсационный фонд возмещения вреда',
                'Размер взноса в компенсационный фонд обеспечения договорных обязательств',
                'Дата вступления в силу решения о приеме в члены',
                'Номер решения о приеме в члены, дата решения о приеме в члены',
                'В отношении объектов капитального строительства (кроме особо опасных, технически сложных и уникальных объектов, объектов использования атомной энергии)',
                'В отношении особо опасных, технически сложных и уникальных объектов капитального строительства (кроме объектов использования атомной энергии)',
                'В отношении объектов использования атомной энергии',
                'Стоимость работ по одному договору подряда (уровень ответственности)',
                'Статус права ВВ',
                'Размер обязательств по договорам подряда с использованием конкурентных способов заключения договоров (уровень ответственности)',
                'Статус права ОДО',
                'Дата оплаты ОДО',
                'Дата доп. взноса ОДО',
                'Дата окончания проверки',
                'Тип проверки',
                'Результат проверки члена СРО',
                'Факты применения мер дисциплинарного воздействия',
                'Предмет договора страхования',
                'Размер страховой суммы',
                'Наименование страховой компании',
                'Местонахождение',
                'Контактные телефоны'
            ]
            for c, v in enumerate(second_row, start=1):
                cell = ws.cell(row=2, column=c, value=v)
                cell.font = Font(bold=True, size=10)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for r in data:
            values_base = [
                r.sro.as_string,
                r.id,
                r.full_description,
                r.short_description,
                r.registration_number,
                r.registry_registration_date,
                r.member_status,
                r.suspension_date,
                r.suspension_reason,
                r.ogrnip,
                r.inn,
                r.phones,
                r.full_address,
                r.director,
                r.accordance_status,
                r.other_information,
                r.members_total_liability,
                r.lialbility_date,
                r.member_right_vv,
                r.member_right_odo,
                r.approved_basis_date,
                r.basis,
                r.is_simple,
                r.is_extremely_dangerous,
                r.is_nuclear,
                r.responsibility_level_vv,
                r.right_status_vv,
                r.responsibility_level_odo,
                r.right_status_odo,
                r.odo_compensation_fund_date,
                r.odo_responsibility_level_date,
            ]
            ws.append(values_base)
            last_row = ws.max_row

            for i_check, check in enumerate(r.checks):
                values_check = [
                    check.check_date,
                    check.member_check_type,
                    check.member_check_result,
                    check.disciplinary_action
                ]
                for i_c, v in enumerate(values_check):
                    ws.cell(row=last_row+i_check, column=len(values_base)+i_c+1, value=v)

            for i_ins, ins in enumerate(r.insurances):
                values_ins = [
                    ins.object_title,
                    ins.insurance_sum,
                    ins.insurer,
                    ins.place,
                    ins.phone,
                ]
                for i_c, v in enumerate(values_ins):
                    ws.cell(row=last_row+i_ins, column=len(values_base)+4+i_c+1, value=v)

            max_row_height = len(r.checks) if len(r.checks) > len(r.insurances) else len(r.insurances)
            # print(max_row_height)
            if max_row_height > 0:
                for i_c in range(1, len(values_base)+1):
                    # print(f'start_row={last_row} | end_row={last_row+max_row_height-1} | column={i_c}')
                    try:
                        ws.merge_cells(start_row=last_row, start_column=i_c,
                                       end_row=last_row+max_row_height-1, end_column=i_c)
                    except Exception as e:
                        logging.info(r)
                        raise e
                    # logging.info(e, exc_info=True)

        wb.save(filepath)
        logging.info(f'{filepath} was written')
