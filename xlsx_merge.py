import asyncio
import logging
from datetime import datetime
from pathlib import Path
from io import BytesIO
from copy import copy
import tqdm
from multiprocessing import cpu_count, Process

import aiofiles
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import DATE_FORMAT
from src.my_logging import get_logger
from src.scrappers.nopriz import ScraperNopriz
from src.scrappers.nostroy import ScraperNostroy
from src.schemas import FiltersNostroy, FiltersNopriz

DATE_FROM = '01.01.1900'
# DATE_FROM = '13.06.2024'
DATE_TO = datetime.now().date().strftime(DATE_FORMAT)
date_from = datetime.strptime(DATE_FROM, DATE_FORMAT)
date_to = datetime.strptime(DATE_TO, DATE_FORMAT)


def get_workbook(filepath: Path) -> Workbook:
    if filepath.exists():
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
    return wb


def append_to_excel(wb_main: Workbook, xlsx_bytes: bytes) -> None:
    ws_main = wb_main.worksheets[0]
    wb_cart = load_workbook(BytesIO(xlsx_bytes))
    ws_cart = wb_cart.worksheets[0]
    last_row = ws_main.max_row - 1 if ws_main.max_row == 1 else ws_main.max_row - 3
    row_slice = 2 if ws_main.max_row > 2 else 0

    for i, row in enumerate(ws_cart.rows):
        if i >= row_slice:
            for cell in row:
                new_row = cell.row

                if cell.column >= 2:
                    if cell.row == 3:
                        continue
                    elif cell.row > 3:
                        new_row = cell.row - 1

                try:
                    new_cell = ws_main.cell(row=last_row + new_row, column=cell.column, value=cell.value)
                except AttributeError:
                    new_cell = ws_main.cell(row=last_row + new_row, column=cell.column, value=None)
                    # print(row)
                    # for l in ws_cart.values:
                    #     print(l)
                    # raise

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    # if cell.column != 1 and cell.row != ws_cart.max_row:
                    new_cell.border = copy(cell.border)
                    # new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    # new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

    for mc in ws_cart.merged_cells.ranges:
        if mc.min_row >= row_slice:
            min_row = mc.min_row
            max_row = mc.max_row

            if mc.min_row == 3 and mc.min_col == 1:
                max_row = mc.max_row - 1
            elif mc.min_row >= 4 and mc.min_col >= 2:
                min_row = mc.min_row - 1
                max_row = mc.max_row - 1

            ws_main.merge_cells(
                start_row=last_row + min_row,
                start_column=mc.min_col,
                end_row=last_row + max_row,
                end_column=mc.max_col
            )


async def _run_scrapper(filepath: Path, wb: Workbook, ids: list[int]):
    async with ScraperNostroy(date_format=DATE_FORMAT, date_from=date_from, date_to=date_to) as scrapper:
        tasks_download = []
        for n, id_ in enumerate(ids):
            tasks_download.append(scrapper.download_xlsx_cart(id_))

            if len(tasks_download) == 20 or id_ == ids[-1]:
                results = await asyncio.gather(*tasks_download)
                print(f'Downloaded bunch of files')

                for xlsx_bytes in results:
                    append_to_excel(wb_main=wb, xlsx_bytes=xlsx_bytes)

                tasks_download = []
                print(f'{n+1}/{len(ids)} Appended data to worksheet')

            if n % 100 == 0:
                print('saving...file')
                wb.save(filepath)
        wb.save(filepath)


def run_scrapper(filepath: Path, wb: Workbook, ids: list[int]) -> None:
    asyncio.run(_run_scrapper(filepath=filepath, wb=wb, ids=ids))


async def main():
    proxy_url = ''

    async with ScraperNostroy(date_format=DATE_FORMAT, date_from=date_from, date_to=date_to) as scrapper:
        filters = FiltersNostroy(member_status=1, sro_enabled=True)
        ids = await scrapper.get_ids(filters=filters, use_cached=True)
        ids = ids[50000:]

        count_processes = 12
        count_ids_per_task = len(ids) // count_processes
        ids_for_tasks = [ids[i:i + count_ids_per_task] for i in range(0, len(ids), count_ids_per_task)]

        for i in range(count_processes):
            filepath = Path(scrapper.get_filename() + f'_{i+1}').with_suffix('.xlsx')
            wb = get_workbook(filepath=filepath)
            p = Process(target=run_scrapper, args=(filepath, wb, ids_for_tasks[i]))
            p.start()


if __name__ == '__main__':
    get_logger(Path(__file__).stem + '.log')

    try:
        asyncio.run(main())
    except Exception as ex:
        logging.error(ex, exc_info=True)
        exit(1)
